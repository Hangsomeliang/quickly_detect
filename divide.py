from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy.datasets import electrocardiogram
from scipy.signal import find_peaks
from scipy.sparse import spdiags
from scipy.linalg import cholesky
import math
import openpyxl
import tkinter as tk
from tkinter import filedialog
import os
# 文件选择ui
root = tk.Tk()
root.withdraw()
f_path = filedialog.askopenfilename(title='请选择XRF检测源文件',filetypes=[("表格","xlsx")])
dirStr, ext = os.path.splitext(f_path)
filename = dirStr.split("/")[-1]
num = filename.split("#")[0]
print(num)
#打开检测结果
test_workbook = load_workbook(f_path)
#打开汇总结果
result_workbook = load_workbook("divide.xlsx")

#检测结果导入
for i in range(1,6+1):
    test_worksheet = test_workbook.get_sheet_by_name(f'{i}')
    result_worksheet = result_workbook.get_sheet_by_name(f'{i}')
    for row in range(test_worksheet.max_row):
        # print(test_worksheet[f'B{row+1}'].value)
        result_worksheet.cell(row = row+1,column=int(num)*2-1).value = test_worksheet.cell(row = row+1,column=1).value
        result_worksheet.cell(row = row+1,column=int(num)*2).value = test_worksheet.cell(row = row+1,column=2).value
#保存汇总结果
result_workbook.save("divide.xlsx")
