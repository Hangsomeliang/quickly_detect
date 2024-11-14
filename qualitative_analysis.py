from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy.datasets import electrocardiogram
from scipy.signal import find_peaks
from scipy.sparse import spdiags
from scipy.linalg import cholesky
import math
import tkinter as tk
from tkinter import filedialog
import openpyxl
import os
from Binary_search import binary_search

#打开特征X射线数据库（XRAYLIB）
xraylib_book = load_workbook("选峰表.xlsx")
#打开6种不同检测参数的数据库
##1号参数
xraylib_sheet1 = xraylib_book['1']
xraylib1_theta = []
xraylib1_element = []
for row in xraylib_sheet1.iter_rows():
    theta = row[8].value
    element = row[1].value
    if theta == None:
        continue
    xraylib1_theta.append(theta)
    xraylib1_element.append(element)
##2号参数
xraylib_sheet2 = xraylib_book['2']
xraylib2_theta = []
xraylib2_element = []
for row in xraylib_sheet2.iter_rows():
    theta = row[8].value
    element = row[1].value
    if theta == None:
        continue
    xraylib2_theta.append(theta)
    xraylib2_element.append(element)
##3号参数
xraylib_sheet3 = xraylib_book['3']
xraylib3_theta = []
xraylib3_element = []
for row in xraylib_sheet3.iter_rows():
    theta = row[8].value
    element = row[1].value
    if theta == None:
        continue
    xraylib3_theta.append(theta)
    xraylib3_element.append(element)
##4号参数
xraylib_sheet4 = xraylib_book['4']
xraylib4_theta = []
xraylib4_element = []
for row in xraylib_sheet4.iter_rows():
    theta = row[8].value
    element = row[1].value
    if theta == None:
        continue
    xraylib4_theta.append(theta)
    xraylib4_element.append(element)
##5号参数
xraylib_sheet5 = xraylib_book['5']
xraylib5_theta = []
xraylib5_element = []
for row in xraylib_sheet5.iter_rows():
    theta = row[8].value
    element = row[1].value
    if theta == None:
        continue
    xraylib5_theta.append(theta)
    xraylib5_element.append(element)
##6号参数
xraylib_sheet6 = xraylib_book['6']
xraylib6_theta = []
xraylib6_element = []
for row in xraylib_sheet6.iter_rows():
    theta = row[8].value
    element = row[1].value
    if theta == None:
        continue
    xraylib6_theta.append(theta)
    xraylib6_element.append(element)
#选择并打开检测源文件
root = tk.Tk()
root.withdraw()
f_path = filedialog.askopenfilename(title='请选择检测文件',filetypes=[("Excel",".xls .xlsx")])
dirStr, ext = os.path.splitext(f_path)
filename = dirStr.split("/")[-1]
print(filename)
#打开检测结果
test_workbook = load_workbook(f_path)
#分析结果保存文件夹选择ui
result_save_path = filedialog.askdirectory(title='请选择分析结果存储路径')
os.chdir(result_save_path) #打开指定目录
try:
    os.makedirs(filename) #创建结果存储文件夹
except:
    None
#创建分析结果存储表格
result_workbook = openpyxl.Workbook()
result_worksheet = result_workbook.active #打开默认工作表
result_number = 1
for i in range(1,6+1):
    #创建检测结果数组
    datax=[]
    datay=[]
    
    #读取检测结果sheet
    test_sheet = test_workbook.get_sheet_by_name(f'{i}')
    for row in test_sheet.iter_rows():
        row_data = row[0].value
        datax.append(row_data)
        row_data = row[1].value
        datay.append(row_data)
    datax = np.array(datax)
    datay = np.array(datay)
    datay_mean= np.mean(datay)
    peaks, _ = find_peaks(datay,width=2,height=datay_mean)
    plt.subplot(2,3,i)
    plt.plot(datax[peaks],datay[peaks],"x")
    plt.axhline(y=datay_mean,color="r",linestyle = "--")
    plt.plot(datax,np.zeros_like(datax),"--",color="gray")
    plt.plot(datax,datay)
    for j in range(len(peaks)):
        print(i)
        print(datax[peaks][j])
        if i==1:
            result = binary_search(xraylib1_theta,datax[peaks][j])
            if(result==None):
                continue
            if(abs(xraylib1_theta[result]-datax[peaks][j])>3):
                continue
            result_worksheet[f'A{result_number}'] = xraylib1_element[result]
        elif i==2:
            result = binary_search(xraylib2_theta,datax[peaks][j])
            if(result==None):
                continue
            if(abs(xraylib2_theta[result]-datax[peaks][j])>3):
                continue
            result_worksheet[f'A{result_number}'] = xraylib2_element[result]
        elif i==3:
            result = binary_search(xraylib3_theta,datax[peaks][j])
            if(result==None):
                continue
            if(abs(xraylib3_theta[result]-datax[peaks][j])>3):
                continue
            result_worksheet[f'A{result_number}'] = xraylib3_element[result]
        elif i==4:
            result = binary_search(xraylib4_theta,datax[peaks][j])
            if(result==None):
                continue
            if(abs(xraylib4_theta[result]-datax[peaks][j])>3):
                continue
            result_worksheet[f'A{result_number}'] = xraylib4_element[result]
        elif i==5:
            result = binary_search(xraylib5_theta,datax[peaks][j])
            if(result==None):
                continue
            if(abs(xraylib5_theta[result]-datax[peaks][j])>3):
                continue
            result_worksheet[f'A{result_number}'] = xraylib5_element[result]
        elif i==6:
            result = binary_search(xraylib6_theta,datax[peaks][j])
            if(result==None):
                continue
            if(abs(xraylib6_theta[result]-datax[peaks][j])>3):
                continue
            result_worksheet[f'A{result_number}'] = xraylib6_element[result]
        result_worksheet[f'B{result_number}'] = datax[peaks][j]
        result_worksheet[f'C{result_number}'] = datay[peaks][j]
        result_worksheet[f'D{result_number}'] = i
        result_number = result_number + 1
result_workbook.save(result_save_path+'/'+filename+'/'+filename+'_result'+'.xlsx')
plt.show()
